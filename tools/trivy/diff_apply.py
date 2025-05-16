"""
Module for applying Trivy diff changes to POAM Excel files.
"""
from pathlib import Path
import json
from typing import Dict, Any, List
from datetime import datetime
import shutil
import openpyxl

def create_updateable_copy(file_path: Path) -> Path:
    """Create a timestamped backup copy of the Excel file."""
    timestamp = datetime.now().strftime('%Y%m%d-%H%M%S')
    backup_path = file_path.parent / f"{file_path.stem}-diff-applied-{timestamp}{file_path.suffix}"
    shutil.copy2(file_path, backup_path)
    return backup_path

def dict_to_row(data: Dict[str, Any]) -> Dict[str, Any]:
    """Convert a dictionary to row format."""
    excel_mapping = {
        "poam_id": "POAM ID",
        "controls": "Controls",
        "weakness_name": "Weakness Name",
        "weakness_description": "Weakness Description",
        "weakness_detector_source": "Weakness Detector Source",
        "weakness_source_identifier": "Weakness Source Identifier",
        "asset_identifier": "Asset Identifier",
        "point_of_contact": "Point of Contact",
        "resources_required": "Resources Required",
        "overall_remediation_plan": "Overall Remediation Plan",
        "original_detection_date": "Original Detection Date",
        "scheduled_completion_date": "Scheduled Completion Date",
        "planned_milestones": "Planned Milestones",
        "milestone_changes": "Milestone Changes",
        "status_date": "Status Date",
        "vendor_dependency": "Vendor Dependency",
        "last_vendor_check_in_date": "Last Vendor Check-in Date",
        "vendor_dependent_product_name": "Vendor Dependent Product Name",
        "original_risk_rating": "Original Risk Rating",
        "adjusted_risk_rating": "Adjusted Risk Rating",
        "risk_adjustment": "Risk Adjustment",
        "false_positive": "False Positive",
        "operational_requirement": "Operational Requirement",
        "deviation_rationale": "Deviation Rationale",
        "supporting_documents": "Supporting Documents",
        "comments": "Comments",
        "auto_approve": "Auto Approve",
        "binding_operational_directive_22_01_tracking": "Binding Operational Directive 22-01 Tracking",
        "binding_operational_directive_22_01_due_date": "Binding Operational Directive 22-01 Due Date",
        "cve": "CVE",
        "service_name": "Service Name"
    }
    return {excel_mapping[k]: v for k, v in data.items() if k in excel_mapping}

def apply_diff(poam_file: Path, diff_json: Dict[str, Any]) -> None:
    """
    Apply diff changes to a POAM Excel file.
    
    Args:
        poam_file: Path to the POAM Excel file
        diff_json: Dictionary containing diff changes
    """
    # Create backup copy
    backup_file = create_updateable_copy(poam_file)
    
    try:
        # Load workbook from backup copy
        wb = openpyxl.load_workbook(backup_file)
        
        # Get sheets
        if "Open POA&M Items" not in wb.sheetnames:
            raise ValueError('Excel file must contain "Open POA&M Items" sheet')
        open_sheet = wb["Open POA&M Items"]
        
        # Get or create closed sheet
        if "Closed POA&M Items" not in wb.sheetnames:
            raise ValueError('Excel file must contain "Open POA&M Items" sheet')
        else:
            closed_sheet = wb["Closed POA&M Items"]

        # Get column indices from header row (row 5)
        header_row = 5
        open_headers = {cell.value: cell.column for cell in open_sheet[header_row]}
        
        # Handle new POAMs - add to open sheet
        if diff_json.get("new_poams"):
            for new_poam in diff_json["new_poams"]:
                row_data = dict_to_row(new_poam["poam"])
                # Add row at the end
                next_row = open_sheet.max_row + 1
                for header, value in row_data.items():
                    if header in open_headers:
                        open_sheet.cell(row=next_row, column=open_headers[header], value=value)
        
        # Handle reopened POAMs - move from closed to open
        if diff_json.get("reopen_poams"):
            reopen_ids = {p["poam_id"] for p in diff_json["reopen_poams"]}
            poam_id_col = next(col for header, col in open_headers.items() if header == "POAM ID")
            
            # Find and move rows
            rows_to_delete = []
            for row in range(header_row + 1, closed_sheet.max_row + 1):
                poam_id = closed_sheet.cell(row=row, column=poam_id_col).value
                if poam_id in reopen_ids:
                    # Copy row to open sheet
                    next_row = open_sheet.max_row + 1
                    for col in range(1, closed_sheet.max_column + 1):
                        open_sheet.cell(row=next_row, column=col, value=closed_sheet.cell(row=row, column=col).value)
                    rows_to_delete.append(row)
            
            # Delete moved rows from closed sheet (in reverse order to maintain indices)
            for row in sorted(rows_to_delete, reverse=True):
                closed_sheet.delete_rows(row)
        
        # Handle closed POAMs - move from open to closed
        if diff_json.get("close_poams"):
            close_ids = set(diff_json["close_poams"])
            poam_id_col = next(col for header, col in open_headers.items() if header == "POAM ID")
            
            # Find and move rows
            rows_to_delete = []
            for row in range(header_row + 1, open_sheet.max_row + 1):
                poam_id = open_sheet.cell(row=row, column=poam_id_col).value
                if poam_id in close_ids:
                    # Copy row to closed sheet
                    next_row = closed_sheet.max_row + 1
                    for col in range(1, open_sheet.max_column + 1):
                        closed_sheet.cell(row=next_row, column=col, value=open_sheet.cell(row=row, column=col).value)
                    rows_to_delete.append(row)
            
            # Delete moved rows from open sheet (in reverse order to maintain indices)
            for row in sorted(rows_to_delete, reverse=True):
                open_sheet.delete_rows(row)
        
        # Save changes to the backup file
        wb.save(backup_file)
        
    except Exception as e:
        # If anything goes wrong, leave the backup file for inspection
        raise type(e)(f"Error applying diff changes. Backup saved as {backup_file}. Error: {str(e)}") from e

def apply_diff_from_files(poam_file: Path, diff_file: Path) -> None:
    """
    Apply diff changes from a JSON file to a POAM Excel file.
    
    Args:
        poam_file: Path to the POAM Excel file
        diff_file: Path to the JSON diff file
    """
    with open(diff_file, 'r') as f:
        diff_json = json.load(f)
    apply_diff(poam_file, diff_json) 